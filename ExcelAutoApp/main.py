import tkinter as tk 
from tkinter import ttk

from sklearn import tree
import openpyxl

def load_data():
    path="people.xlsx"
    global work_book
    work_book=openpyxl.load_workbook(path)
    global work_sheet
    work_sheet=work_book.active
    list_values=list(work_sheet.values)
    for col_name in list_values[0]:
        treeview.heading(col_name,text=col_name)
    for row in list_values[1:]:
        treeview.insert("",tk.END,values=row)

    # wb=openpyxl.load_workbook("ExcelAutoApp/people.xlsx")
    # ws=wb.active
    # for row in ws.iter_rows(min_row=2,values_only=True):
    #     treeview.insert("",tk.END,values=row)
def insert_row():
    #insert data into excel file
    name=name_entry.get()
    age=int(age_spinbox.get())
    subscription=status_combo.get()
    employment_status="Employed" if a.get() else "Unemployed"
    new_row=[name,age,subscription,employment_status]
    work_sheet.append(new_row)
    work_book.save("people.xlsx")
    #Insert data into treeview
    treeview.insert("",tk.END,values=new_row)
    #clear entry fields
    name_entry.delete(0,"end")
    name_entry.insert(0,"Name")
    age_spinbox.delete(0,"end")
    age_spinbox.insert(0,"Age")
    status_combo.current(0)
    a.set(False)

def toggle_mode():

    if mode_switch.instate(["selected"]):
        style.theme_use("forest-light")
    else:
        style.theme_use("forest-dark")

root=tk.Tk()
style=ttk.Style(root)
root.tk.call("source","Forest-ttk-theme/forest-light.tcl")
root.tk.call("source","Forest-ttk-theme/forest-dark.tcl")
style.theme_use('forest-dark')

combo_list=["Subscribed","Not Subscribe", "Other"]

frame=ttk.Frame(root)
frame.pack()

widgets_frame=ttk.LabelFrame(frame,text="Insert Row")
widgets_frame.grid(row=0,column=0,padx=20,pady=10)

name_entry=ttk.Entry(widgets_frame)
name_entry.insert(0,"Name")
name_entry.bind("<FocusIn>",lambda e: name_entry.delete(0,"end"))
name_entry.grid(row=0,column=0 ,sticky="ew", padx=5,pady=(0,5))

age_spinbox=ttk.Spinbox(widgets_frame,from_=18,to=100)
age_spinbox.insert(0,"Age")
age_spinbox.grid(row=1,column=0,sticky="ew",padx=5,pady=5)

status_combo=ttk.Combobox(widgets_frame,values=combo_list)
status_combo.current(0)
status_combo.grid(row=2,column=0,sticky="ew",padx=5,pady=5)

a=tk.BooleanVar()
chackbutton=ttk.Checkbutton(widgets_frame,text="Employed",variable=a)
chackbutton.grid(row=3,column=0,sticky="nsew",padx=5,pady=5)

button=ttk.Button(widgets_frame,text="Insert",command=insert_row)
button.grid(row=4,column=0,sticky="nsew",padx=5,pady=5)

separtor=ttk.Separator(widgets_frame)
separtor.grid(row=5,column=0,sticky="ew",padx=(20,10),pady=5)

mode_switch=ttk.Checkbutton(widgets_frame,text="Mode",style="Switch",command=toggle_mode)
mode_switch.grid(row=6,column=0,sticky="nsew",padx=5,pady=10)

treeFrame=ttk.Frame(frame)
treeFrame.grid(row=0,column=1,pady=10)
treescroll=ttk.Scrollbar(treeFrame)
treescroll.pack(side="right",fill="y")

cols=["Name","Age","Subscription","Employment"]
treeview=ttk.Treeview(treeFrame,show="headings",columns=cols,height=13)
treeview.column("Name",width=100)
treeview.column("Age",width=50)
treeview.column("Subscription",width=100)
treeview.column("Employment",width=100)
treeview.pack(side="left",fill="y")
treeview.configure(yscrollcommand=treescroll.set)
treescroll.configure(command=treeview.yview)

load_data()









root.mainloop()
